import os
import re
import subprocess
import sys
from pathlib import Path
from openpyxl import Workbook
from typing import Dict

def extract_error_positions(directory):
    missed_report = []
    missed_report_message = []
    false_positive = []
    false_positive_message = []
    error_sums_after = []
    error_sums_before = []
    error_all = []
    for filename in os.listdir(directory):
        if filename.endswith('.log'):
            filepath = os.path.join(directory, filename)
            error_positions_after = []
            error_mess_after = []
            error_positions_before = []
            error_mess_before = []
            with open(filepath, 'r', encoding='gbk') as file:
                content = file.readlines()
                for line in content:
                    matches = re.findall(r'\[(\d+),\d+\]', line)
                    for match in matches:
                        error_positions_after.append(int(match)-3)
                        error_mess_after.append(line[46:])
                        print(int(match), filename, line)
                        print()
                    
            try:
                org_path = os.path.join(r'C:\Users\Asfdf\Desktop\annoation\0307logs', filename)
                with open(org_path, 'r', encoding='gbk') as file:
                    content = file.readlines()
                for line in content:
                    matches = re.findall(r'\[(\d+),\d+\]', line)
                    for match in matches:
                        error_positions_before.append(int(match))
                        error_mess_before.append(line[47:])
                        print(int(match), filename, line)
                        print()
                mis = 0
                fal = 0
                # before
                # for i in range(len(error_mess_after)):
                #     if error_positions_after[i]-3 not in error_positions_before:
                #         mis = mis + 1
                #         missed_report_message.append([filename, mis, error_mess_after[i]])
                #     else:
                #         error_all.append([filename, error_mess_after[i]])
                # for i in range(len(error_positions_before)):
                #     if error_positions_before[i]+3 not in error_positions_after:
                #         fal = fal + 1
                #         false_positive_message.append([filename, fal, error_mess_before[i]])
                for i in range(len(error_mess_after)):
                    if error_positions_after[i] not in error_positions_before:
                        mis = mis + 1
                        missed_report_message.append([filename, mis, error_mess_after[i]])
                    else:
                        error_all.append([filename, error_mess_after[i]])
                for i in range(len(error_positions_before)):
                    if error_positions_before[i] not in error_positions_after:
                        fal = fal + 1
                        false_positive_message.append([filename, fal, error_mess_before[i]])
                missed_report.append([filename, mis])
                false_positive.append([filename, fal])
                error_sums_after.append(len(error_positions_after))
                error_sums_before.append(len(error_positions_before))
            except:
                print(f"没有找到{filename}对应的log文件")
    
    return missed_report, missed_report_message, false_positive, false_positive_message, error_sums_before, error_sums_after, error_all
def write_data(root_dir, missed_report, missed_report_message, false_positive, false_positive_message, error_sums_before, error_sums_after, error_all):
    os.chdir(root_dir)
    wb = Workbook()
    ws = wb.active
    ws.title = "漏报消除情况(召回率)"
    ws.append(["样例名称", "漏报数目","被正确报告数目", "总数量"])
    for i in range(len(missed_report)):
        ws.append([missed_report[i][0], missed_report[i][1], error_sums_after[i]-missed_report[i][1], error_sums_after[i]])
    ws = wb.create_sheet("误报消除情况(准确率)")
    ws.append(["样例名称", "误报数目", "正确报告数目", "总数量"])
    for i in range(len(false_positive)):
        ws.append([false_positive[i][0], false_positive[i][1], error_sums_before[i]-false_positive[i][1], error_sums_before[i]])
    ws = wb.create_sheet("漏报消除详细信息")
    ws.append(["样例名称", "消除漏报序号", "报错详细信息"])
    for i in range(len(missed_report_message)):
        ws.append([missed_report_message[i][0], missed_report_message[i][1], missed_report_message[i][2]])
    ws = wb.create_sheet("误报消除详细信息")
    ws.append(["样例名称", "消除误报序号", "报错详细信息"])
    for i in range(len(false_positive_message)):
        ws.append([false_positive_message[i][0], false_positive_message[i][1], false_positive_message[i][2]])
    ws = wb.create_sheet("报错交集")
    ws.append(["样例名称" , "报错详细信息"])
    for i in range(len(error_all)):
        ws.append([error_all[i][0], error_all[i][1]])
    wb.save("analyze_result_0307_org.xlsx")

directory = r'C:\Users\Asfdf\Desktop\annoation\0307logs_after'
missed_report, missed_report_message, false_positive, false_positive_message, error_sums_before, error_sums_after, error_all = extract_error_positions(directory)
write_data(r'C:\Users\Asfdf\Desktop\annoation', missed_report, missed_report_message, false_positive, false_positive_message, error_sums_before, error_sums_after, error_all)

